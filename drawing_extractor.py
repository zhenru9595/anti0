import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import json
import threading
from pathlib import Path
import io
import time
import base64

from google import genai
import openai
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import fitz  # PyMuPDF
import re

try:
    import pytesseract
    HAS_TESSERACT = True
except ImportError:
    HAS_TESSERACT = False

try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False

# ── 색상 팔레트 ────────────────────────────────────────────────
BG_DARK    = "#12131a"
BG_PANEL   = "#1c1e2b"
BG_INPUT   = "#23263a"
ACCENT     = "#4f8ef7"
ACCENT2    = "#2ecc71"
TEXT_MAIN  = "#e8eaf0"
TEXT_SUB   = "#7880a0"
ROW_ODD    = "#1c2033"
ROW_EVEN   = "#1a1c2a"
BORDER_COL = "#2d3050"

CONFIG_FILE = Path(__file__).parent / ".drawing_extractor_config.json"

SUPPORTED_EXT = {".pdf", ".png", ".jpg", ".jpeg"}
GEMINI_MODELS = ["gemini-2.0-flash", "gemini-2.0-flash-lite", "gemini-1.5-flash-latest"]

def load_config() -> dict:
    try:
        if CONFIG_FILE.exists():
            return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    except Exception:
        pass
    # 하위 호환성 위해 예전 버전의 api_key도 읽음
    try:
        if CONFIG_FILE.exists():
            old = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            if "api_key" in old and "gemini_key" not in old:
                return {"gemini_key": old["api_key"], "provider": "Gemini"}
    except:
        pass
    return {"provider": "Gemini", "gemini_key": "", "openai_key": ""}

def save_config(config: dict):
    try:
        CONFIG_FILE.write_text(json.dumps(config, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass

def pil_to_base64(img: Image.Image) -> str:
    buffered = io.BytesIO()
    if img.mode != 'RGB':
        img = img.convert('RGB')
    img.save(buffered, format="JPEG", quality=85)
    return base64.b64encode(buffered.getvalue()).decode('utf-8')

class DrawingExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("도면 정보 추출기  ·  Drawing Extractor (Gemini & OpenAI 지원)")
        self.root.geometry("960x780")
        self.root.minsize(800, 650)
        self.root.configure(bg=BG_DARK)

        self.config = load_config()
        self.file_path: str | None = None
        self.extracted_data: dict | None = None
        self.is_busy = False
        self._session = 0

        self._apply_style()
        self._build_ui()
        self._update_provider_ui()

    def _apply_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        background=ROW_EVEN, foreground=TEXT_MAIN,
                        rowheight=30, fieldbackground=ROW_EVEN,
                        font=("Segoe UI", 10), borderwidth=0)
        style.configure("Treeview.Heading",
                        background=BG_INPUT, foreground=ACCENT,
                        font=("Segoe UI", 10, "bold"), relief="flat", borderwidth=0)
        style.map("Treeview",
                  background=[("selected", "#2a3560")],
                  foreground=[("selected", TEXT_MAIN)])
        style.configure("Vertical.TScrollbar",
                        background=BG_INPUT, troughcolor=BG_PANEL,
                        arrowcolor=TEXT_SUB, borderwidth=0)

    def _build_ui(self):
        # ── 헤더 ─────────────────────────────────────────────
        hdr = tk.Frame(self.root, bg=BG_DARK)
        hdr.pack(fill=tk.X, pady=(18, 6), padx=24)
        tk.Label(hdr, text="🔧", font=("Segoe UI Emoji", 22), bg=BG_DARK, fg=ACCENT).pack(side=tk.LEFT)
        tk.Label(hdr, text=" 도면 정보 추출기", font=("Segoe UI", 20, "bold"), bg=BG_DARK, fg=TEXT_MAIN).pack(side=tk.LEFT)
        tk.Label(hdr, text="  PDF · PNG · JPG → Excel", font=("Segoe UI", 11), bg=BG_DARK, fg=TEXT_SUB).pack(side=tk.LEFT, pady=(4, 0))

        # ── API 설정 ──────────────────────────────────────────
        self._section_label("AI Provider & API 설정")
        api_card = self._card()

        # 제공자 선택 라디오버튼
        prov_row = tk.Frame(api_card, bg=BG_PANEL)
        prov_row.pack(fill=tk.X, padx=14, pady=(10, 0))
        tk.Label(prov_row, text="AI 모델", width=8, font=("Segoe UI", 10, "bold"), fg=TEXT_MAIN, bg=BG_PANEL, anchor="w").pack(side=tk.LEFT)
        
        self.provider_var = tk.StringVar(value=self.config.get("provider", "Gemini"))
        for val in ["Gemini", "OpenAI", "Local OCR (무료)"]:
            rb = tk.Radiobutton(prov_row, text=val, variable=self.provider_var, value=val,
                                bg=BG_PANEL, fg=TEXT_MAIN, selectcolor=BG_INPUT,
                                activebackground=BG_PANEL, activeforeground=ACCENT, cursor="hand2",
                                command=self._update_provider_ui, font=("Segoe UI", 10))
            rb.pack(side=tk.LEFT, padx=(0, 15))

        # API 키 입력
        row = tk.Frame(api_card, bg=BG_PANEL)
        row.pack(fill=tk.X, padx=14, pady=8)
        tk.Label(row, text="API Key", width=8, font=("Segoe UI", 10), fg=TEXT_SUB, bg=BG_PANEL, anchor="w").pack(side=tk.LEFT)

        self.api_var = tk.StringVar()
        self.api_entry = tk.Entry(row, textvariable=self.api_var, show="•", width=50,
                                  bg=BG_INPUT, fg=TEXT_MAIN, insertbackground=ACCENT,
                                  relief=tk.FLAT, font=("Segoe UI", 10))
        self.api_entry.pack(side=tk.LEFT, padx=(8, 8), ipady=5)

        self.show_var = tk.BooleanVar()
        tk.Checkbutton(row, text="표시", variable=self.show_var, command=self._toggle_key,
                       bg=BG_PANEL, fg=TEXT_SUB, selectcolor=BG_INPUT,
                       activebackground=BG_PANEL, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        tk.Button(row, text="키 저장", bg="#2d3050", fg=ACCENT,
                  activebackground="#3a3f6a", activeforeground=ACCENT,
                  font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0,
                  padx=10, pady=3, cursor="hand2",
                  command=self._save_api_key).pack(side=tk.LEFT, padx=(8, 0))

        # 안내 링크
        self.guide_lbl = tk.Label(api_card, text="", font=("Segoe UI", 9), fg="#5577cc", bg=BG_PANEL, cursor="hand2")
        self.guide_lbl.pack(anchor=tk.W, padx=14, pady=(0, 4))
        self.guide_lbl.bind("<Button-1>", lambda _: self._open_url(self.guide_url))

        self.key_status_var = tk.StringVar(value="")
        tk.Label(api_card, textvariable=self.key_status_var, font=("Segoe UI", 9), fg=ACCENT2, bg=BG_PANEL).pack(anchor=tk.W, padx=14, pady=(0, 6))

        # ── 파일 선택 (+ 드래그 앤 드롭) ─────────────────────
        self._section_label("도면 파일 선택")
        file_card = self._card()

        self.drop_frame = tk.Frame(file_card, bg=BG_INPUT, height=100, cursor="hand2")
        self.drop_frame.pack(fill=tk.X, padx=14, pady=(10, 6))
        self.drop_frame.pack_propagate(False)

        dnd_hint = "📂   클릭하거나 파일을 이곳에 드래그하세요\nPDF · PNG · JPG · JPEG 지원"
        self.drop_lbl = tk.Label(self.drop_frame, text=dnd_hint, font=("Segoe UI", 11), fg=ACCENT, bg=BG_INPUT, justify=tk.CENTER, cursor="hand2")
        self.drop_lbl.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        for w in (self.drop_frame, self.drop_lbl):
            w.bind("<Button-1>", lambda _: self._browse())
            w.bind("<Enter>", lambda _: self.drop_frame.config(bg="#2a2d42"))
            w.bind("<Leave>", lambda _: self.drop_frame.config(bg=BG_INPUT))

        if DND_AVAILABLE:
            self.drop_frame.drop_target_register(DND_FILES)
            self.drop_frame.dnd_bind("<<Drop>>", self._on_drop)
            self.drop_lbl.drop_target_register(DND_FILES)
            self.drop_lbl.dnd_bind("<<Drop>>", self._on_drop)

        self.file_lbl = tk.Label(file_card, text="선택된 파일 없음", font=("Segoe UI", 9), fg=TEXT_SUB, bg=BG_PANEL)
        self.file_lbl.pack(pady=(0, 8))

        # ── 버튼 바 ──────────────────────────────────────────
        btn_bar = tk.Frame(self.root, bg=BG_DARK)
        btn_bar.pack(fill=tk.X, padx=24, pady=(4, 0))

        self.extract_btn = self._btn(btn_bar, "🔍  도면 분석 시작", ACCENT, self._start_extraction)
        self.extract_btn.pack(side=tk.LEFT)
        self.stop_btn = self._btn(btn_bar, "🛑  작업 정지", "#e74c3c", self._stop_extraction, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=(10, 0))
        self.save_btn = self._btn(btn_bar, "💾  엑셀로 저장", ACCENT2, self._save_excel, state=tk.DISABLED)
        self.save_btn.pack(side=tk.LEFT, padx=(10, 0))

        # ── 상태 바 ──────────────────────────────────────────
        self.status_var = tk.StringVar(value="준비 완료  —  AI 모델을 선택하고 도면 파일을 업로드하세요.")
        tk.Label(self.root, textvariable=self.status_var, font=("Segoe UI", 9), fg=TEXT_SUB, bg=BG_DARK, anchor=tk.W).pack(fill=tk.X, padx=26, pady=(6, 2))

        # ── 결과 테이블 ───────────────────────────────────────
        self._section_label("추출 결과")
        res_card = self._card(expand=True)

        cols = ("항목", "추출 값")
        self.tree = ttk.Treeview(res_card, columns=cols, show="headings", height=12)
        self.tree.heading("항목", text="  항목")
        self.tree.heading("추출 값", text="추출 값")
        self.tree.column("항목", width=160, minwidth=120, anchor=tk.CENTER)
        self.tree.column("추출 값", width=700, minwidth=300, anchor=tk.W)

        vsb = ttk.Scrollbar(res_card, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0), pady=10)
        vsb.pack(side=tk.RIGHT, fill=tk.Y, pady=10, padx=(0, 8))

        self.tree.tag_configure("odd",  background=ROW_ODD)
        self.tree.tag_configure("even", background=ROW_EVEN)

    def _update_provider_ui(self):
        prov = self.provider_var.get()
        if prov == "Gemini":
            key = self.config.get("gemini_key", "")
            url = "https://aistudio.google.com/app/apikey"
            text = f"🔑  Gemini API 키 발급받기 → {url} (무료 티어 15RPM/1500RPD)"
            self.api_entry.config(state=tk.NORMAL)
        elif prov == "OpenAI":
            key = self.config.get("openai_key", "")
            url = "https://platform.openai.com/api-keys"
            text = f"🔑  OpenAI API 키 발급받기 → {url} (유료, 매우 향상된 성능 보장)"
            self.api_entry.config(state=tk.NORMAL)
        else: # Local OCR
            key = "API 키가 필요 없습니다 (비용 완전 무료)"
            url = "https://github.com/UB-Mannheim/tesseract/wiki"
            text = f"⚙️  Tesseract-OCR 설치가 필요합니다 → {url}"
            self.api_entry.config(state=tk.DISABLED)
        
        self.api_var.set(key)
        self.guide_lbl.config(text=text)
        self.guide_url = url
        if prov != "Local OCR (무료)" and key:
            self.key_status_var.set(f"✅  저장된 {prov} API 키를 불러왔습니다.")
        else:
            self.key_status_var.set("")

    # ── 헬퍼 위젯 ────────────────────────────────────────────
    def _section_label(self, text):
        tk.Label(self.root, text=text, font=("Segoe UI", 10, "bold"), fg=TEXT_SUB, bg=BG_DARK).pack(anchor=tk.W, padx=26, pady=(10, 2))

    def _card(self, expand=False):
        f = tk.Frame(self.root, bg=BG_PANEL, highlightthickness=1, highlightbackground=BORDER_COL)
        f.pack(fill=tk.BOTH if expand else tk.X, expand=expand, padx=24, pady=2)
        return f

    @staticmethod
    def _btn(parent, text, color, cmd, state=tk.NORMAL):
        return tk.Button(parent, text=text, bg=color, fg="white", activebackground=color, activeforeground="white",
                         font=("Segoe UI", 11, "bold"), relief=tk.FLAT, bd=0, padx=18, pady=8, cursor="hand2", command=cmd, state=state)

    def _toggle_key(self):
        self.api_entry.config(show="" if self.show_var.get() else "•")

    def _save_api_key(self):
        key = self.api_var.get().strip()
        prov = self.provider_var.get()
        if not key:
            messagebox.showwarning("경고", "API 키를 입력한 뒤 저장해주세요.")
            return
        
        self.config["provider"] = prov
        if prov == "Gemini":
            self.config["gemini_key"] = key
        else:
            self.config["openai_key"] = key
            
        save_config(self.config)
        self.key_status_var.set(f"✅  {prov} API 키가 저장되었습니다.")

    @staticmethod
    def _open_url(url):
        import webbrowser
        webbrowser.open(url)

    def _set_file(self, path: str):
        ext = Path(path).suffix.lower()
        if ext not in SUPPORTED_EXT:
            messagebox.showwarning("지원하지 않는 형식", f"'{ext}' 파일은 지원되지 않습니다.\nPDF, PNG, JPG, JPEG 파일만 사용 가능합니다.")
            return
        self.file_path = path
        name = Path(path).name
        self.file_lbl.config(text=f"✅  {name}", fg=ACCENT2)
        self.drop_frame.config(bg=BG_INPUT)
        self.status_var.set(f"파일 선택됨: {name}")

    def _browse(self):
        path = filedialog.askopenfilename(
            title="도면 파일 선택",
            filetypes=[("지원 형식", "*.pdf *.png *.jpg *.jpeg"), ("PDF", "*.pdf"), ("이미지", "*.png *.jpg *.jpeg"), ("모두", "*.*")]
        )
        if path:
            self._set_file(path)

    def _on_drop(self, event):
        raw = event.data.strip()
        if raw.startswith("{") and raw.endswith("}"):
            raw = raw[1:-1]
        path = raw.split("} {")[0] if "} {" in raw else raw.split()[0] if " " in raw and not Path(raw).exists() else raw
        self._set_file(path)

    def _stop_extraction(self):
        """진행 중인 모든 분석 및 대기열(타이머) 강제 종료"""
        self._session += 1
        self.is_busy = False
        self.extract_btn.config(state=tk.NORMAL, text="🔍  도면 분석 시작")
        self.stop_btn.config(state=tk.DISABLED)
        self.status_var.set("🛑  사용자가 작업을 강제 중지했습니다.")

    def _start_extraction(self):
        if self.is_busy:
            return
        if not self.file_path:
            messagebox.showwarning("파일 없음", "도면 파일을 먼저 선택해주세요.")
            return
        api_key = self.api_var.get().strip()
        prov = self.provider_var.get()
        if not api_key and prov != "Local OCR (무료)":
            messagebox.showwarning("API 키 없음", f"{prov} API 키를 입력해주세요.")
            return

        self.is_busy = True
        self._session += 1
        cur_session = self._session
        self.extract_btn.config(state=tk.DISABLED, text=f"⏳  {prov} 분석 중...")
        self.stop_btn.config(state=tk.NORMAL)
        self.save_btn.config(state=tk.DISABLED)
        self.status_var.set(f"분석 중입니다... (최대 1분 소요)")
        threading.Thread(target=self._extract, args=(api_key, prov, 0, cur_session), daemon=True).start()

    # ── AI 분석 (백그라운드 스레드) ───────────────────────────
    def _extract(self, api_key: str, provider: str, retry_count: int = 0, session: int = 0):
        MAX_RETRY = 2
        WAIT_SEC  = 60
        if session != self._session:
            return
            
        try:
            image = self._load_image()
            prompt = (
                "아래 도면 이미지를 분석하여 공학/제조 정보를 추출해주세요.\n"
                "반드시 아래 JSON 키만 사용하여 응답하세요. 값이 없으면 '정보 없음'으로 표기.\n"
                "단위는 값과 함께 표기 (예: 100mm, R5, 1.2kg).\n\n"
                "{\n"
                '  "너비": "...",\n'
                '  "길이": "...",\n'
                '  "높이": "...",\n'
                '  "반경": "...",\n'
                '  "무게": "...",\n'
                '  "재질": "...",\n'
                '  "표면처리": "...",\n'
                '  "열처리": "...",\n'
                '  "기타 공정": "...",\n'
                '  "기타 표기": "..."\n'
                "}\n\nJSON만 출력하고 다른 텍스트는 포함하지 마세요."
            )

            text = ""
            if provider == "Gemini":
                client = genai.Client(api_key=api_key)
                last_err = None
                for model_name in GEMINI_MODELS:
                    try:
                        response = client.models.generate_content(
                            model=model_name,
                            contents=[prompt, image],
                        )
                        text = response.text.strip()
                        break
                    except Exception as e:
                        last_err = e
                        if "404" in str(e) or "NOT_FOUND" in str(e):
                            continue
                        raise
                else:
                    raise last_err

            elif provider == "OpenAI":
                b64_img = pil_to_base64(image)
                client = openai.OpenAI(api_key=api_key)
                response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt},
                                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_img}"}},
                            ],
                        }
                    ],
                    max_tokens=2000,
                )
                text = response.choices[0].message.content.strip()

            elif provider == "Local OCR (무료)":
                if not HAS_TESSERACT:
                    raise RuntimeError("pytesseract 패키지가 설치되지 않았습니다. 'pip install pytesseract' 실행 필요.")
                
                tess_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
                if os.path.exists(tess_path):
                    pytesseract.pytesseract.tesseract_cmd = tess_path
                    
                try:
                    ocr_text = pytesseract.image_to_string(image, lang='eng+kor')
                except Exception as e:
                    if "tesseract is not installed" in str(e).lower() or "not found" in str(e).lower():
                        raise RuntimeError("Tesseract 프로그램이 설치되지 않았습니다.\nhttps://github.com/UB-Mannheim/tesseract/wiki 에서 윈도우 설치파일을 다운로드해주세요.")
                    raise

                # 단순 룰 기반 필터링
                data = {
                    "너비": "정보 없음", "길이": "정보 없음", "높이": "정보 없음",
                    "반경": "정보 없음", "무게": "정보 없음", "재질": "정보 없음", 
                    "기타 표기": ocr_text.strip()
                }
                
                # 치수 추출 (예: 100mm, M4, R5)
                dims = set(re.findall(r'(\d+(?:\.\d+)?[mM]{2})', ocr_text))
                radiuses = set(re.findall(r'[Rr]\d+', ocr_text))
                threads = set(re.findall(r'[Mm]\d+', ocr_text))
                weight = set(re.findall(r'(\d+(?:\.\d+)?[kK]?[gG])', ocr_text))

                if dims: data["기타 공정"] = "추출된 치수: " + ", ".join(dims)
                if threads: data["너비"] = ", ".join(threads)
                if radiuses: data["반경"] = ", ".join(radiuses)
                if weight: data["무게"] = ", ".join(weight)

                self.extracted_data = data
                self.root.after(0, self._show_results, data)
                return

            # 마크다운 코드 블록 제거
            if "```" in text:
                for part in text.split("```"):
                    p = part.strip().lstrip("json").strip()
                    if p.startswith("{"):
                        text = p
                        break

            data = json.loads(text)
            self.extracted_data = data
            self.root.after(0, self._show_results, data)

        except json.JSONDecodeError:
            raw = locals().get("text", "")
            self.root.after(0, self._on_error, f"JSON 파싱 실패.\nAI 응답:\n{raw[:400]}")
        except Exception as exc:
            err = str(exc)
            
            # OpenAI 에러 처리
            if "insufficient_quota" in err or "429" in err or "RESOURCE_EXHAUSTED" in err:
                if retry_count < MAX_RETRY and provider == "Gemini": # OpenAI는 429일시 돈 부족/한도이므로 바로 에러 리턴 권장
                    def countdown(secs_left):
                        if session != self._session:  # 도중 정지 시 타이머 취소
                            return
                        if secs_left > 0:
                            self.root.after(0, self.status_var.set, f"⏳  API 한도 초과 — {secs_left}초 후 자동 재시도합니다... (시도 {retry_count+1}/{MAX_RETRY})")
                            threading.Timer(1, countdown, args=(secs_left - 1,)).start()
                        else:
                            self.root.after(0, self.status_var.set, "🔄  재시도 중...")
                            threading.Thread(target=self._extract, args=(api_key, provider, retry_count + 1, session), daemon=True).start()
                    countdown(WAIT_SEC)
                else:
                    self.root.after(0, self._on_error,
                        f"API 할당량(크레딧) 초과 또는 트래픽 제한 (429)\n\n"
                        f"• {provider} 계정에 결제 수단이 등록되어 있거나 한도가 넉넉한지 확인하세요.\n"
                        f"• OpenAI의 경우 선불 크레딧 충전(Billing)이 필요합니다.\n"
                        f"• 증상이 지속되면 새로운 API 키를 발급받으세요.")
            elif "API_KEY_INVALID" in err or "API key not valid" in err or "Incorrect API key" in err:
                self.root.after(0, self._on_error, f"{provider} API 키가 유효하지 않습니다. 키를 확인해주세요.")
            else:
                self.root.after(0, self._on_error, err)

    def _load_image(self) -> Image.Image:
        ext = Path(self.file_path).suffix.lower()
        if ext == ".pdf":
            doc = fitz.open(self.file_path)
            pix = doc[0].get_pixmap(matrix=fitz.Matrix(2, 2))
            doc.close()
            return Image.open(io.BytesIO(pix.tobytes("png")))
        return Image.open(self.file_path)

    def _show_results(self, data: dict):
        for row in self.tree.get_children():
            self.tree.delete(row)
        fields = ["너비", "길이", "높이", "반경", "무게", "재질", "표면처리", "열처리", "기타 공정", "기타 표기"]
        for i, key in enumerate(fields):
            tag = "odd" if i % 2 else "even"
            self.tree.insert("", tk.END, values=(key, data.get(key, "정보 없음")), tags=(tag,))
        self.is_busy = False
        self.extract_btn.config(state=tk.NORMAL, text="🔍  도면 분석 시작")
        self.stop_btn.config(state=tk.DISABLED)
        self.save_btn.config(state=tk.NORMAL)
        self.status_var.set("✅  분석 완료!  '엑셀로 저장' 버튼으로 결과를 저장하세요.")

    def _on_error(self, msg: str):
        self.is_busy = False
        self.extract_btn.config(state=tk.NORMAL, text="🔍  도면 분석 시작")
        self.stop_btn.config(state=tk.DISABLED)
        self.status_var.set("❌  오류 발생")
        messagebox.showerror("분석 오류", f"분석 중 오류가 발생했습니다:\n\n{msg}")

    def _save_excel(self):
        if not self.extracted_data:
            return
        save_path = filedialog.asksaveasfilename(
            title="엑셀 파일 저장", defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")], initialfile="도면_추출_결과.xlsx"
        )
        if not save_path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "도면 정보"
            thin = Border(
                left=Side(style="thin", color="D0D5E8"), right=Side(style="thin", color="D0D5E8"),
                top=Side(style="thin", color="D0D5E8"), bottom=Side(style="thin", color="D0D5E8"),
            )
            center = Alignment(horizontal="center", vertical="center")
            left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.merge_cells("A1:B1")
            ws["A1"] = "도면 정보 추출 결과"
            ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
            ws["A1"].fill = PatternFill("solid", fgColor="4F8EF7")
            ws["A1"].alignment = center
            ws.row_dimensions[1].height = 34
            ws.merge_cells("A2:B2")
            ws["A2"] = f"원본 파일: {Path(self.file_path).name}"
            ws["A2"].font = Font(size=9, color="888888")
            ws["A2"].alignment = center
            ws.row_dimensions[2].height = 18

            for col, label in zip(("A", "B"), ("항목", "추출 값")):
                c = ws[f"{col}3"]
                c.value = label
                c.font = Font(bold=True, color="FFFFFF", size=11)
                c.fill = PatternFill("solid", fgColor="2C3560")
                c.alignment = center
                c.border = thin
            ws.row_dimensions[3].height = 26

            fields = ["너비", "길이", "높이", "반경", "무게", "재질", "표면처리", "열처리", "기타 공정", "기타 표기"]
            for i, key in enumerate(fields, start=4):
                bg = "F5F7FF" if i % 2 == 0 else "FFFFFF"
                ws[f"A{i}"] = key
                ws[f"B{i}"] = self.extracted_data.get(key, "")
                ws[f"A{i}"].font = Font(bold=True, color="1A1A2E", size=11)
                ws[f"B{i}"].font = Font(color="1A1A2E", size=11)
                for col in ("A", "B"):
                    ws[f"{col}{i}"].fill = PatternFill("solid", fgColor=bg)
                    ws[f"{col}{i}"].border = thin
                ws[f"A{i}"].alignment = center
                ws[f"B{i}"].alignment = left
                ws.row_dimensions[i].height = 22

            ws.column_dimensions["A"].width = 16
            ws.column_dimensions["B"].width = 62
            wb.save(save_path)
            self.status_var.set(f"✅  저장 완료: {save_path}")
            if messagebox.askyesno("저장 완료", f"엑셀 저장 완료!\n{save_path}\n\n파일을 바로 여시겠습니까?"):
                os.startfile(save_path)
        except Exception as exc:
            messagebox.showerror("저장 오류", str(exc))

if __name__ == "__main__":
    if DND_AVAILABLE:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    app = DrawingExtractorApp(root)
    root.mainloop()
